from openpyxl import load_workbook

from config import EXCEL_FILE, SHEET_NAME

BATCH_COLUMN = 3       # Column C
SHIPMENT_COLUMN = 6    # Column F


def update_shipment_excel(shipment_dict):

    wb = load_workbook(EXCEL_FILE)

    ws = wb[SHEET_NAME]

    updated = 0

    for row in range(2, ws.max_row + 1):

        batch = ws.cell(row=row, column=BATCH_COLUMN).value

        if batch is None:
            continue

        batch = str(batch).strip()

        if batch in shipment_dict:

            ws.cell(
                row=row,
                column=SHIPMENT_COLUMN
            ).value = shipment_dict[batch]

            updated += 1

    wb.save(EXCEL_FILE)
    wb.close()

    print(f"[OK] Updated {updated} Shipment.")