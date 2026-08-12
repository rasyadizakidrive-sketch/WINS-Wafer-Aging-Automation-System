import sys
import os

sys.path.append(
    os.path.dirname(
        os.path.dirname(
            os.path.abspath(__file__)
        )
    )
)

from openpyxl import load_workbook

import config

# ==========================================
# COLUMN
# ==========================================

BATCH_COLUMN = 3      # Column C
GR_DATE_COLUMN = 7    # Column G


# ==========================================
# GET BATCH WITHOUT GR DATE
# ==========================================

def get_batch_without_gr_date():

    wb = load_workbook(config.EXCEL_FILE)

    ws = wb[config.SHEET_NAME]

    batch_list = []

    for row in range(2, ws.max_row + 1):

        batch = ws.cell(
            row=row,
            column=BATCH_COLUMN
        ).value

        gr_date = ws.cell(
            row=row,
            column=GR_DATE_COLUMN
        ).value

        if batch is None:
            continue

        batch = str(batch).strip()

        # Only Batch start with 000
        if not batch.startswith("000"):
            continue

        # Only empty GR Date
        if gr_date is None or str(gr_date).strip() == "":

            batch_list.append(batch)

    wb.close()

    return batch_list


# ==========================================
# TEST
# ==========================================

if __name__ == "__main__":

    batches = get_batch_without_gr_date()

    print(f"Found {len(batches)} Batch\n")

    for batch in batches:

        print(batch)