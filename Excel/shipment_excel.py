import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
import config

BATCH_COLUMN = 3      # C
SHIPMENT_COLUMN = 6   # F


def get_batch_without_shipment():

    wb = load_workbook(config.EXCEL_FILE)
    ws = wb[config.SHEET_NAME]

    batch_list = []

    for row in range(2, ws.max_row + 1):

        batch = ws.cell(row=row, column=BATCH_COLUMN).value
        shipment = ws.cell(row=row, column=SHIPMENT_COLUMN).value

        if batch is None:
            continue

        batch = str(batch).strip()

        # Hanya ambil batch yang bermula dengan 000
        if not batch.startswith("000"):
            continue

        # Hanya ambil yang Shipment masih kosong
        if shipment is None or str(shipment).strip() == "":
            batch_list.append(batch)
    wb.close()

    return batch_list


if __name__ == "__main__":

    batches = get_batch_without_shipment()

    print(f"Found {len(batches)} Batch\n")

    for b in batches:
        print(b)